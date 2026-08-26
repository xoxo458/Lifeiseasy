"""Write the extracted statement into the required Excel layout.

Layout (matching the target format):
  row 1-2  account header block
  row 3    styled, filtered, frozen column headers
  row 4+   one transaction per row, dates as real dates and amounts as numbers
A second sheet records the validation result so the output is auditable.
"""

from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Statement
from .validate import Report

HEADER_ROW = 3
FIRST_DATA_ROW = HEADER_ROW + 1

# (header, attribute, width, number format, horizontal alignment)
COLUMNS = [
    ("Tran. Date",     "tran_date",      12, "d-mmm-yyyy", "left"),
    ("Effect Date",    "effect_date",    12, "dd-mmm-yy",  "left"),
    ("Tran. Br.",      "branch",          9, "0",          "left"),
    ("Description",    "description",    62, None,         "left"),
    ("Remitter Name",  "remitter_name",  22, None,         "left"),
    ("Remitter IBAN",  "remitter_iban",  26, None,         "left"),
    ("Remitter Bank",  "remitter_bank",  16, None,         "left"),
    ("Chq / Ref No",   "ref_no",         14, "@",          "left"),
    ("Debit",          "debit",          16, "#,##0.00",   "right"),
    ("Credit",         "credit",         16, "#,##0.00",   "right"),
    ("Balance",        "balance",        18, "#,##0.00",   "right"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=11)
_THIN = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def write(statement: Statement, out_path: str, report: Report | None = None) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"

    _write_header_block(ws, statement)
    _write_columns(ws)
    _write_rows(ws, statement)
    _finish_sheet(ws, statement)

    if report is not None:
        _write_validation_sheet(wb, statement, report)

    wb.save(out_path)
    return out_path


def _write_header_block(ws, statement) -> None:
    meta = statement.meta
    period = ""
    if meta.period_from and meta.period_to:
        period = f"{meta.period_from:%d-%b-%Y} to {meta.period_to:%d-%b-%Y}"

    left = " | ".join(p for p in [meta.account_title, meta.branch] if p)
    right = " | ".join(
        p for p in [
            f"A/C No: {meta.account_no}" if meta.account_no else "",
            f"IBAN: {meta.iban}" if meta.iban else "",
            f"{meta.account_type}/{meta.currency}" if meta.account_type or meta.currency else "",
        ] if p
    )
    ws.cell(row=1, column=1, value=left or "Account Statement").font = TITLE_FONT
    ws.cell(row=1, column=5, value=right)
    ws.cell(row=2, column=1, value=f"Statement Period: {period}" if period else "")
    if meta.opening_balance is not None:
        ws.cell(row=2, column=9, value="Opening Balance:").font = TITLE_FONT
        cell = ws.cell(row=2, column=11, value=_number(meta.opening_balance))
        cell.number_format = "#,##0.00"


def _write_columns(ws) -> None:
    for index, (header, _attr, width, _fmt, _align) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[HEADER_ROW].height = 20


def _write_rows(ws, statement) -> None:
    for offset, txn in enumerate(statement.transactions):
        row = FIRST_DATA_ROW + offset
        for index, (_header, attr, _width, fmt, align) in enumerate(COLUMNS, start=1):
            value = getattr(txn, attr)
            cell = ws.cell(row=row, column=index, value=_excel_value(attr, value))
            if fmt:
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal=align, vertical="top")
            cell.border = CELL_BORDER


def _excel_value(attr: str, value):
    if value is None or value == "":
        return None
    if attr == "branch":
        return int(value) if str(value).isdigit() else str(value)
    if attr == "ref_no":
        return str(value)  # text format keeps leading zeros
    if isinstance(value, Decimal):
        return _number(value)
    return value


def _number(value: Decimal) -> float:
    return float(value)


def _finish_sheet(ws, statement) -> None:
    last_row = FIRST_DATA_ROW + len(statement.transactions) - 1
    if last_row < FIRST_DATA_ROW:
        last_row = HEADER_ROW
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{last_row}"
    ws.freeze_panes = f"A{FIRST_DATA_ROW}"


def _write_validation_sheet(wb, statement, report: Report) -> None:
    ws = wb.create_sheet("Validation")
    meta, totals = statement.meta, statement.totals

    rows = [
        ("Source file", statement.source_file),
        ("Extraction engine", statement.engine),
        ("Transactions extracted", report.checked_rows),
        ("Balance chain", "OK" if report.balance_chain_ok else "FAILED"),
        ("Footer totals", "OK" if report.totals_ok else "FAILED"),
        ("Errors", len(report.errors)),
        ("Warnings", len(report.warnings)),
        ("", ""),
        ("Opening balance (printed)", _opt(meta.opening_balance)),
        ("Closing balance (printed)", _opt(totals.closing_balance)),
        ("Available balance (printed)", _opt(totals.available_balance)),
        ("Total DR transactions (printed)", totals.total_dr_count),
        ("Total CR transactions (printed)", totals.total_cr_count),
        ("Sum of DR transactions (printed)", _opt(totals.sum_dr)),
        ("Sum of CR transactions (printed)", _opt(totals.sum_cr)),
    ]
    for index, (label, value) in enumerate(rows, start=1):
        ws.cell(row=index, column=1, value=label).font = TITLE_FONT if label else Font()
        ws.cell(row=index, column=2, value=value)

    start = len(rows) + 2
    for column, header in enumerate(("Severity", "Row", "Check", "Detail"), start=1):
        cell = ws.cell(row=start, column=column, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for offset, issue in enumerate(report.issues, start=1):
        ws.cell(row=start + offset, column=1, value=issue.severity)
        ws.cell(row=start + offset, column=2, value=issue.row)
        ws.cell(row=start + offset, column=3, value=issue.check)
        ws.cell(row=start + offset, column=4, value=issue.detail)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 90


def _opt(value):
    return float(value) if isinstance(value, Decimal) else value
