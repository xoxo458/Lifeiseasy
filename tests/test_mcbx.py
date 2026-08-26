"""Tests for the MCB statement converter.

The end-to-end tests run against a generated digital statement (tests/make_fixture.py)
whose rows mirror the real scanned sample, including a row that wraps across a
page break. The vision engine's network call is not exercised; its payload
handling is tested against a recorded response shape.
"""

from __future__ import annotations

import sys
from decimal import Decimal
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from mcbx import engine_vision, excel, parse, validate
from mcbx.models import Statement, StatementMeta, StatementTotals, Transaction, parse_amount, parse_date
from mcbx.parse import RawLine


# --- primitives -------------------------------------------------------------

@pytest.mark.parametrize(
    "raw,expected",
    [
        ("01-AUG-26", (2026, 8, 1)),
        ("17-APR-20", (2020, 4, 17)),
        ("9-dec-99", (1999, 12, 9)),
        ("01-AUG-2026", (2026, 8, 1)),
    ],
)
def test_parse_date(raw, expected):
    got = parse_date(raw)
    assert (got.year, got.month, got.day) == expected


@pytest.mark.parametrize("raw", ["", None, "junk", "32-XXX-26", "2026-08-01"])
def test_parse_date_rejects(raw):
    assert parse_date(raw) is None


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("1,144,333.79", Decimal("1144333.79")),
        ("0.16", Decimal("0.16")),
        ("58,080,277.00", Decimal("58080277.00")),
        ("(50.00)", Decimal("-50.00")),
        ("", None),
        ("-", None),
        (None, None),
        ("abc", None),
    ],
)
def test_parse_amount(raw, expected):
    assert parse_amount(raw) == expected


# --- stitching --------------------------------------------------------------

def line(page=1, **cells):
    return RawLine(cells=cells, page=page)


def test_wrapped_row_is_stitched_without_spaces():
    """MCB breaks mid-word; the target format rejoins with no separator."""
    lines = [
        line(tran_date="01-AUG-26", effect_date="01-AUG-26", branch="1160",
             description="FEDERAL EXCISE", debit="0.16", balance="10,509,604.71"),
        line(description="DUTY TAX/"),
    ]
    (txn,) = parse.stitch(lines)
    assert txn.description == "FEDERAL EXCISEDUTY TAX/"
    assert txn.debit == Decimal("0.16")
    assert txn.credit is None


def test_wrap_join_space_option():
    lines = [
        line(tran_date="01-AUG-26", description="FEDERAL EXCISE", debit="0.16", balance="1.00"),
        line(description="DUTY TAX/"),
    ]
    (txn,) = parse.stitch(lines, wrap_join="space")
    assert txn.description == "FEDERAL EXCISE DUTY TAX/"


def test_wrapped_iban_is_reassembled():
    lines = [
        line(tran_date="01-AUG-26", description="P2P RECEIVING VIA",
             remitter_name="JUNAID", remitter_iban="PK54MEZN001",
             credit="94,000.00", balance="10,603,604.71"),
        line(description="RAAST-OFFUS/", remitter_name="AHMAD BAIG", remitter_iban="252010732301"),
        line(description="Purpose: -", remitter_iban="2"),
    ]
    (txn,) = parse.stitch(lines)
    assert txn.remitter_iban == "PK54MEZN0012520107323012"
    assert txn.remitter_name == "JUNAIDAHMAD BAIG"


def test_row_split_across_page_break_stays_one_transaction():
    lines = [
        line(page=1, tran_date="05-AUG-26", description="CMD TRANSFER",
             credit="10,000,000.00", balance="21,217,661.79"),
        line(page=1, description="CREDIT/FUNDSTRF-"),
        line(page=2, description="Limited"),  # tail printed on the next page
    ]
    (txn,) = parse.stitch(lines)
    assert txn.description == "CMD TRANSFERCREDIT/FUNDSTRF-Limited"


def test_footer_and_header_noise_is_dropped():
    lines = [
        line(tran_date="01-AUG-26", description="MONTHLY BUNDLE", debit="1.00", balance="10,509,604.87"),
        line(description="Total DR Transactions", debit="32"),
        line(description="Sum of CR Transactions", debit="77,919,000.00"),
        line(tran_date="Tran. Date", description="Transaction Details"),
        line(description="Page: 1 of 4"),
        line(description="User ID: mcb1160"),
    ]
    txns = parse.stitch(lines)
    assert len(txns) == 1
    assert txns[0].description == "MONTHLY BUNDLE"


def test_leading_orphan_continuation_is_not_a_transaction():
    """A tail whose parent row was never captured must not become a dateless row."""
    lines = [
        line(page=2, description="Limited (PayDirect)"),
        line(page=2, tran_date="06-AUG-26", description="FEDERAL EXCISE", debit="112.00", balance="1.00"),
    ]
    txns = parse.stitch(lines)
    assert len(txns) == 1
    assert txns[0].description == "FEDERAL EXCISE"


def test_branch_leading_zeros_are_stripped():
    lines = [line(tran_date="05-AUG-26", branch="0069", description="X", debit="1.00", balance="1.00")]
    assert parse.stitch(lines)[0].branch == "69"


def test_dedupe_keeps_genuinely_repeated_rows():
    """Consecutive identical charges are real; only exact re-reports are dropped."""
    a = Transaction(tran_date=parse_date("01-AUG-26"), description="SETTLEMENT",
                    debit=Decimal("28647.64"), balance=Decimal("10574957.07"))
    b = Transaction(tran_date=parse_date("01-AUG-26"), description="SETTLEMENT",
                    debit=Decimal("28647.64"), balance=Decimal("10546309.43"))
    assert len(parse.dedupe([a, b])) == 2
    assert len(parse.dedupe([a, a])) == 1


# --- validation -------------------------------------------------------------

def _statement(rows, opening="100.00", **totals):
    st = Statement(meta=StatementMeta(opening_balance=Decimal(opening)))
    st.transactions = rows
    st.totals = StatementTotals(**totals)
    return st


def test_balance_chain_detects_a_wrong_figure():
    rows = [
        Transaction(tran_date=parse_date("01-AUG-26"), description="A",
                    debit=Decimal("10.00"), balance=Decimal("90.00")),
        Transaction(tran_date=parse_date("02-AUG-26"), description="B",
                    credit=Decimal("5.00"), balance=Decimal("99.00")),  # should be 95.00
    ]
    report = validate.validate(_statement(rows))
    assert not report.balance_chain_ok
    assert any(i.check == "balance-chain" and i.row == 2 for i in report.errors)


def test_balance_chain_accepts_a_correct_statement():
    rows = [
        Transaction(tran_date=parse_date("01-AUG-26"), description="A",
                    debit=Decimal("10.00"), balance=Decimal("90.00")),
        Transaction(tran_date=parse_date("02-AUG-26"), description="B",
                    credit=Decimal("5.00"), balance=Decimal("95.00")),
    ]
    report = validate.validate(
        _statement(rows, total_dr_count=1, total_cr_count=1,
                   sum_dr=Decimal("10.00"), sum_cr=Decimal("5.00"),
                   closing_balance=Decimal("95.00"))
    )
    assert report.ok and report.balance_chain_ok and report.totals_ok


def test_footer_totals_mismatch_is_an_error():
    rows = [Transaction(tran_date=parse_date("01-AUG-26"), description="A",
                        debit=Decimal("10.00"), balance=Decimal("90.00"))]
    report = validate.validate(
        _statement(rows, total_dr_count=2, total_cr_count=0,
                   sum_dr=Decimal("10.00"), sum_cr=Decimal("0.00"),
                   closing_balance=Decimal("90.00"))
    )
    assert not report.totals_ok
    assert any(i.check == "count-dr" for i in report.errors)


def test_row_with_both_debit_and_credit_is_an_error():
    rows = [Transaction(tran_date=parse_date("01-AUG-26"), description="A",
                        debit=Decimal("1.00"), credit=Decimal("1.00"), balance=Decimal("100.00"))]
    report = validate.validate(_statement(rows))
    assert any(i.check == "amount" for i in report.errors)


# --- vision engine payload handling (no network) ----------------------------

RECORDED_PAYLOAD = {
    "transactions": [
        {
            "continues_from_previous": False,
            "tran_date": "01-AUG-26", "effect_date": "01-AUG-26", "branch": "5398",
            "description_lines": ["P2P RECEIVING VIA", "RAAST-OFFUS/", "Purpose: -"],
            "remitter_name_lines": ["JUNAID", "AHMAD BAIG"],
            "remitter_iban_lines": ["PK54MEZN001", "252010732301", "2"],
            "remitter_bank_lines": [], "ref_no": "",
            "debit": "", "credit": "94,000.00", "balance": "10,603,604.71",
        },
        {
            "continues_from_previous": False,
            "tran_date": "18-AUG-26", "effect_date": "18-AUG-26", "branch": "1771",
            "description_lines": ["OUTWARD CHEQUE", "CLEARING CREDIT/"],
            "remitter_name_lines": [], "remitter_iban_lines": [], "remitter_bank_lines": [],
            "ref_no": "0000014203",
            "debit": "", "credit": "1,500,000.00", "balance": "2,948,513.79",
        },
    ],
    "meta": {
        "account_title": "AMAZON MARKETING (SMC-PRIVATE) LIMITED", "address": "",
        "account_no": "1175809501009469", "iban": "PK90MUCB1175809501009469",
        "account_type": "BUS", "currency": "PKR", "branch": "1160-ISLAMABAD JINAH SUPER MKT",
        "opened_on": "17-APR-20", "period_from": "01-AUG-26", "period_to": "19-AUG-26",
        "statement_datetime": "Aug 19, 2026 09:39:13 AM", "opening_balance": "10,509,605.87",
    },
    "totals": {
        "total_dr_count": "32", "total_cr_count": "20",
        "sum_dr": "85,480,092.08", "sum_cr": "77,919,000.00",
        "available_balance": "2,948,513.79", "closing_balance": "2,948,513.79",
    },
}


def test_vision_payload_expands_to_stitchable_lines():
    lines = engine_vision._payload_lines(RECORDED_PAYLOAD, first_page=1)
    txns = parse.stitch(lines)
    assert len(txns) == 2
    assert txns[0].description == "P2P RECEIVING VIARAAST-OFFUS/Purpose: -"
    assert txns[0].remitter_iban == "PK54MEZN0012520107323012"
    assert txns[0].credit == Decimal("94000.00")
    assert txns[1].ref_no == "0000014203"


def test_vision_continuation_flag_merges_into_previous_row():
    payload = {
        "transactions": [{
            "continues_from_previous": True,
            "tran_date": "", "effect_date": "", "branch": "",
            "description_lines": ["Limited"],
            "remitter_name_lines": [], "remitter_iban_lines": [], "remitter_bank_lines": [],
            "ref_no": "", "debit": "", "credit": "", "balance": "",
        }],
        "meta": {}, "totals": {},
    }
    first = engine_vision._payload_lines(RECORDED_PAYLOAD, first_page=1)
    second = engine_vision._payload_lines(payload, first_page=2)
    txns = parse.stitch(first + second)
    assert len(txns) == 2
    assert txns[-1].description == "OUTWARD CHEQUECLEARING CREDIT/Limited"


def test_vision_meta_and_totals_merge():
    st = Statement()
    engine_vision._merge_meta(st, RECORDED_PAYLOAD["meta"])
    engine_vision._merge_totals(st, RECORDED_PAYLOAD["totals"])
    assert st.meta.account_no == "1175809501009469"
    assert st.meta.opening_balance == Decimal("10509605.87")
    assert st.totals.total_dr_count == 32
    assert st.totals.sum_cr == Decimal("77919000.00")


# --- end to end -------------------------------------------------------------

@pytest.fixture(scope="module")
def fixture_pdf(tmp_path_factory):
    from make_fixture import build

    return build(str(tmp_path_factory.mktemp("pdf") / "mcb_fixture.pdf"))


@pytest.fixture(scope="module")
def converted(fixture_pdf):
    from mcbx.engine_text import extract

    lines, statement = extract(fixture_pdf)
    statement.transactions = parse.dedupe(parse.stitch(lines))
    return statement, validate.validate(statement)


def test_end_to_end_extracts_every_row_and_reconciles(converted):
    statement, report = converted
    assert len(statement.transactions) == 9
    assert report.balance_chain_ok, validate.summarise(report)
    assert report.totals_ok, validate.summarise(report)
    assert report.ok, validate.summarise(report)


def test_end_to_end_header_block(converted):
    statement, _ = converted
    meta = statement.meta
    assert meta.account_title == "AMAZON MARKETING (SMC-PRIVATE) LIMITED"
    assert meta.account_no == "1175809501009469"
    assert meta.iban == "PK90MUCB1175809501009469"
    assert meta.currency == "PKR"
    assert meta.opening_balance == Decimal("10509605.87")
    assert (meta.period_from.month, meta.period_to.day) == (8, 19)


def test_end_to_end_page_spanning_row(converted):
    """The row split across the page break must come back whole."""
    statement, _ = converted
    inward = [t for t in statement.transactions if t.description.startswith("INWARD CHEQUE")]
    assert len(inward) == 1
    assert inward[0].description == (
        "INWARD CHEQUECLEARINGDEBIT/2066146083 - TMNIFT CUSTOMERCHEQUE:2066146083"
    )
    assert inward[0].debit == Decimal("2500000.00")


def test_end_to_end_excel_layout(converted, tmp_path):
    from openpyxl import load_workbook

    statement, report = converted
    out = tmp_path / "out.xlsx"
    excel.write(statement, str(out), report)

    wb = load_workbook(out)
    ws = wb["Statement"]
    assert [c.value for c in ws[excel.HEADER_ROW]] == [c[0] for c in excel.COLUMNS]
    assert ws.freeze_panes == f"A{excel.FIRST_DATA_ROW}"
    assert ws.auto_filter.ref.startswith(f"A{excel.HEADER_ROW}:")

    first = excel.FIRST_DATA_ROW
    assert ws.cell(row=first, column=1).value.strftime("%d-%b-%Y") == "01-Aug-2026"
    assert ws.cell(row=first, column=1).number_format == "d-mmm-yyyy"
    assert ws.cell(row=first, column=9).value == 1.0          # debit is a number
    assert ws.cell(row=first, column=10).value is None        # blank credit, not 0
    assert ws.cell(row=first, column=11).number_format == "#,##0.00"

    # Branch 0069 becomes 69; a reference number keeps its leading zeros as text.
    branches = [ws.cell(row=first + i, column=3).value for i in range(9)]
    assert 69 in branches
    refs = [ws.cell(row=first + i, column=8).value for i in range(9)]
    assert "0000014203" in refs

    assert "Validation" in wb.sheetnames


def test_cli_writes_a_workbook(fixture_pdf, tmp_path):
    from mcbx.cli import main

    out = tmp_path / "cli.xlsx"
    assert main([fixture_pdf, "-o", str(out), "--strict", "-q"]) == 0
    assert out.is_file()


def test_auth_errors_are_recognised():
    assert engine_vision._is_auth_error(
        Exception("Could not resolve authentication method. Expected one of api_key...")
    )
    assert not engine_vision._is_auth_error(Exception("connection reset by peer"))


def test_excel_handles_a_statement_with_no_rows(tmp_path):
    from openpyxl import load_workbook

    st = Statement(source_file="empty.pdf", engine="text")
    report = validate.validate(st)
    out = tmp_path / "empty.xlsx"
    excel.write(st, str(out), report)
    ws = load_workbook(out)["Statement"]
    assert ws.cell(row=excel.HEADER_ROW, column=1).value == "Tran. Date"


# --- offline OCR engine -----------------------------------------------------

def test_ocr_engine_reads_a_clean_statement(fixture_pdf):
    """Tesseract handles a crisp render; watermarked photos are another matter
    (see the accuracy note in the README)."""
    from mcbx.engine_ocr import extract as extract_ocr, tesseract_available

    if not tesseract_available():
        pytest.skip("tesseract is not installed")

    lines, statement = extract_ocr(fixture_pdf)
    statement.transactions = parse.dedupe(parse.stitch(lines))
    report = validate.validate(statement)

    assert len(statement.transactions) == 9
    assert report.balance_chain_ok, validate.summarise(report)
    assert report.totals_ok, validate.summarise(report)
